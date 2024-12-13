from openpyxl import Workbook


def create_phone_number_excel(phone_number, repetitions, file_name="phone_numbers.xlsx"):
    # Create a new Excel workbook and select the active sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Phone Numbers"

    # Set a header for the phone number column
    ws["A1"] = "Phone Number"

    # Fill the column with the phone number, repeated as many times as specified
    for row in range(2, repetitions + 2):
        ws[f"A{row}"] = phone_number

    # Save the workbook to a file
    wb.save(file_name)
    print(f"Excel file '{file_name}' created with {repetitions} entries of the phone number {phone_number}.")


# Example usage
create_phone_number_excel("7062731309", 100)  # Change the number and repetitions as needed
